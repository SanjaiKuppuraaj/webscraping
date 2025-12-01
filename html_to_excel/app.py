# app/html_to_excel_blueprint.py
from flask import Blueprint, render_template, abort, current_app, Response, request
from bs4 import BeautifulSoup as bs
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from io import BytesIO
import os
from common_code import common_module as cm

html_to_excel_bp = Blueprint('html_to_excel', __name__, template_folder='templates')

html_folder = os.path.join(cm.BASE_DIR_OUTPUTS, 'html_to_excel', 'html_input')


def get_td_text(tds, index):
    return tds[index].get_text(separator=' ', strip=True) if len(tds) > index else None


def convert_html_to_excel_in_memory(filename, subpath=""):
    try:
        file_path = os.path.join(html_folder, subpath, filename) if subpath else os.path.join(html_folder, filename)

        if not os.path.exists(file_path):
            abort(404, description=f"File {filename} not found in folder {subpath}")

        with open(file_path, "r", encoding="utf-8") as f:
            soup = bs(f.read(), "html.parser")

        rows = soup.find_all('tr')
        results = []
        current_case, party_name, address_list = None, [], []

        for row in rows:
            case_status = ""
            fname_lower = filename.lower()
            if "pending" in fname_lower:
                case_status = "Pending"
            elif "disposal" in fname_lower:
                case_status = "Disposal"

            tds = row.find_all('td')
            if not tds:
                continue

            first_line = get_td_text(tds, 0)

            if first_line and re.match(r'^\d+$', first_line.strip()):
                if current_case:
                    current_case['government_party'] = party_name
                    current_case['address_grouped'] = address_list
                    results.append(current_case)

                current_case = {}
                party_name, address_list = [], []
                current_case['sno'] = int(first_line.strip())

                details = get_td_text(tds, 1).replace('Enter OIC PDF', '').split()
                case_no = details[0].split('/')
                current_case['case_type'] = case_no[0]
                current_case['case_no'] = int(case_no[1])
                current_case['case_year'] = int(case_no[2])
                current_case['bench'] = ' '.join(details[1:]).strip().split(' ')[-1]

                parties_html = str(tds[2])
                parsed_parties = bs(parties_html, 'html.parser')
                separate_parties = parsed_parties.find_all('b')
                judge_name = separate_parties[1].find_parent('span').text if len(separate_parties) > 1 else ""

                parties_text = parsed_parties.text.replace('\n', '').replace(judge_name, '').strip()
                parties = parties_text.split('Vs')

                respondent = str(parties[1]) if len(parties) > 1 else ""
                s_d_w_match = re.search(r"S/D/W/.*?(?=District:-|Police Station:|$)", respondent)
                district_match = re.search(r"District:-.*?(?=Police Station:|$)", respondent)
                police_match = re.search(r"Police Station:-.*", respondent)

                s_d_w_part = s_d_w_match.group().strip() if s_d_w_match else ""
                current_case['thru'] = str(s_d_w_part).split('Crime No')[0].strip().replace('        ', '') if s_d_w_part else ''

                district_part = district_match.group().strip() if district_match else ""
                current_case['district'] = str(district_part).split(':-')[1].strip() if ':-' in district_part else ''

                police_part = police_match.group().strip() if police_match else ""
                current_case['police_station'] = str(police_part).split(':-')[1].strip() if ':-' in police_part else ''

                cleaned_respondent = respondent
                for part in [s_d_w_part, district_part, police_part]:
                    if part:
                        cleaned_respondent = cleaned_respondent.replace(part, '')

                cleaned_respondent = cleaned_respondent.split('Crime No')[0]
                current_case['respondent'] = cleaned_respondent.strip(' ,;:-')

                petitioner = parties[0].split(']')
                pet_no = petitioner[0].strip() + ' ]'
                current_case['respondent/petitioner_no'] = pet_no if pet_no else ''

                for k in pet_no:
                    if 'P' in k:
                        current_case['apperingas'] = 'Petitioner'
                    elif 'R' in k:
                        current_case['apperingas'] = 'Respondent'

                current_case['petitioner'] = petitioner[1].strip() if len(petitioner) > 1 else ''

                current_case['case_status'] = case_status

                if pname := get_td_text(tds, 3):
                    party_name.append(pname)
                if addr := get_td_text(tds, 4):
                    address_list.append(addr.strip())

                current_case['department'] = [get_td_text(tds, 5)] if get_td_text(tds, 5) else []
                current_case['reply/rejoinder/return'] = [get_td_text(tds, 6)] if get_td_text(tds, 6) else []
            elif current_case:
                if pname := get_td_text(tds, 0):
                    party_name.append(pname)
                if addr := get_td_text(tds, 1):
                    address_list.append(addr.strip())
                if dept := get_td_text(tds, 2):
                    current_case['department'].append(dept)
                if rejoinder := get_td_text(tds, 3):
                    current_case['reply/rejoinder/return'].append(rejoinder)

        if current_case:
            current_case['government_party'] = party_name
            current_case['address_grouped'] = address_list
            results.append(current_case)

        flattened = []
        for case in results:
            flat_case = case.copy()
            flat_case['government_party'] = '\n'.join(case['government_party']) if isinstance(case['government_party'], list) else case['government_party']
            flat_case['address_grouped'] = '\n'.join([','.join([p.strip() for p in a.split(',') if p.strip()]) for a in case['address_grouped']]) if isinstance(case['address_grouped'], list) else case['address_grouped']
            flat_case['department'] = '\n'.join(case['department']) if isinstance(case['department'], list) else case['department']
            flat_case['reply/rejoinder/return'] = '\n'.join(case['reply/rejoinder/return']) if isinstance(case['reply/rejoinder/return'], list) else case['reply/rejoinder/return']
            flattened.append(flat_case)

        df = pd.DataFrame(flattened)
        wb = Workbook()
        ws = wb.active
        ws.append(df.columns.to_list())
        for row in df.itertuples(index=False):
            ws.append(list(row))

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    except Exception as e:
        current_app.logger.error(f"Error in convert_html_to_excel_in_memory for {filename} in {subpath}: {e}", exc_info=True)
        abort(500, description="Internal Server Error")


@html_to_excel_bp.route("", defaults={"subpath": ""})
@html_to_excel_bp.route("/browse/", defaults={"subpath": ""})
@html_to_excel_bp.route("/browse/<path:subpath>")
def browse_and_convert(subpath):
    try:
        safe_subpath = os.path.normpath(subpath).replace("..", "")
        full_path = os.path.join(html_folder, safe_subpath)

        download_param = request.args.get("download") or ""

        if not os.path.exists(full_path):
            abort(404, description=f"Path '{subpath}' not found")

        if os.path.isdir(full_path):
            entries = os.listdir(full_path)
            dirs = sorted([d for d in entries if os.path.isdir(os.path.join(full_path, d))])

            def extract_year_and_status(filename):
                match = re.search(r'_(\d{4})\.html$', filename)
                year = int(match.group(1)) if match else 0
                status = 0 if 'pending' in filename.lower() else 1
                return (year, -status)

            html_files = [f for f in entries if f.endswith('.html')]
            files = sorted(html_files, key=extract_year_and_status, reverse=True)

            show_download = (download_param == "122.166.101.95")

            return render_template(
                "html_to_excelindex.html",
                subpath=subpath,
                files=files,
                dirs=dirs,
                show_download=show_download,
                download_param=download_param
            )

        else:
            abort(404, description="Invalid access. Please browse folders and download files.")

    except Exception as e:
        current_app.logger.error(f"Error browsing path '{subpath}': {e}", exc_info=True)
        abort(500, description="Failed to browse directory")


@html_to_excel_bp.route("/download/<path:subpath>/<filename>")
def generate_excel(subpath, filename):
    try:
        safe_subpath = os.path.normpath(subpath).replace("..", "")
        safe_filename = os.path.basename(filename)
        file_path = os.path.join(html_folder, safe_subpath, safe_filename)

        if not os.path.exists(file_path):
            abort(404, description=f"File {filename} not found in folder {subpath}")

        excel_data = convert_html_to_excel_in_memory(safe_filename, safe_subpath)
        return Response(
            excel_data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={os.path.splitext(safe_filename)[0]}.xlsx"}
        )

    except Exception as e:
        current_app.logger.error(f"Failed to generate Excel for {filename} in {subpath}: {e}", exc_info=True)
        abort(500, description="Failed to generate Excel file")


@html_to_excel_bp.route("/download_all/<path:subpath>")
def generate_combined_excel(subpath):
    try:
        safe_subpath = os.path.normpath(subpath).replace("..", "")
        folder_path = os.path.join(html_folder, safe_subpath)

        if not os.path.exists(folder_path) or not os.path.isdir(folder_path):
            abort(404, description=f"Folder {subpath} not found or is not a directory")

        all_data = []

        for filename in sorted(os.listdir(folder_path)):
            if filename.endswith(".html"):
                try:
                    excel_bytes = convert_html_to_excel_in_memory(filename, safe_subpath)
                    df = pd.read_excel(BytesIO(excel_bytes))
                    df.insert(0, 'Source File', filename)
                    all_data.append(df)
                except Exception as inner_e:
                    current_app.logger.warning(f"Skipping {filename} due to error: {inner_e}")

        if not all_data:
            abort(400, description="No valid HTML files found to convert.")

        combined_df = pd.concat(all_data, ignore_index=True)

        wb = Workbook()
        ws = wb.active
        ws.append(combined_df.columns.tolist())
        for row in combined_df.itertuples(index=False):
            ws.append(list(row))

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return Response(stream.getvalue(),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment;filename=Combined_{safe_subpath.replace('/', '_')}.xlsx"})

    except Exception as e:
        current_app.logger.error(f"Failed to generate combined Excel for folder {subpath}: {e}", exc_info=True)
        abort(500, description="Failed to generate combined Excel file")
